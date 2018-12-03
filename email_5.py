import poplib

poplib._MAXLINE = 204800
import time
import os
from email.parser import Parser
from email.header import decode_header
from email.utils import parseaddr, parsedate
from openpyxl import Workbook
import re
import datetime
from tenacity import retry, wait_fixed, stop_after_attempt
import xlrd

charset_collect = ['gb2312', 'gb18030', 'gbk', 'iso-8859-2', 'big5', 'iso-8859-6', 'iso-8859-4', 'iso-8859-5',
                   'iso-8859-7', 'utf-16', 'iso-8859-1', 'iso-2022-kr', 'iso-8859-3', 'iso-2022-jp', 'iso-2022-jp',
                   'iso-8859-15', 'iso-8859-9',
                   'iso-8859-8-i', 'iso-8859-8']
address_list = ['tg.gtja.com', 'chinastock.com.cn', 'swhysc.com', 'stocke.com.cn', 'orientsec.com.cn', 'citics.com',
                'cmschina.com.cn', 'gfund.com', 'service.pingan.com', 'crctrust.com', 'cifutures.com.cn', 'htfc.com']


# 解码,解析加密传输的字符串成正常字符串
def decode_str(s):
    try:
        value, charset = decode_header(s)[0]
        if charset:
            value = value.decode(charset)
        return value, charset
    except:
        pass


@retry(wait=wait_fixed(3), stop=stop_after_attempt(3))
# 登录邮箱, 获取邮件字符串源文件,解析出主题/发件人邮箱,发送时间, 是否带有附件
def log_in(host, user, passwd):
    # 开始登录
    pop_conn = poplib.POP3_SSL(host)
    pop_conn.user(user)
    pop_conn.pass_(passwd)
    pop_conn.noop()
    mail_count = len(pop_conn.list()[1])
    # 显示邮箱状态：邮件数量，占用空间
    # print('Messages: %s. Size: %s' % pop_conn.stat())
    # print('测试信息...\n邮箱登陆成功! \n收件箱共有%s封邮件,占用空间%s字节\n' % (mail_count, pop_conn.stat()[1]))
    return mail_count, pop_conn


# 多字符集尝试解码邮件
def decode_email_content(lines):
    counter = 1
    str_lines = ''
    got_charset = ''
    for charset in charset_collect:
        try:
            str_lines = b'\r\n'.join(lines).decode(charset)
            got_charset = charset
            break
        except:
            counter = counter + 1
            continue
    print('在第%s次尝试后, 用%s字符集解码邮件成功' % (counter, got_charset))
    msg = Parser().parsestr(str_lines)
    return msg


def validate_date(date_):
    if parsedate(date_):
        date_tuple = parsedate(date_)
        send_date_str = time.strftime("%Y%m%d", date_tuple)
    else:
        send_date_str = ''.join(
            re.findall('.*?([1,2]{1}[0,9]{1}[0-9]{2})[/-]([0,1]{1}[0-9]{1})[/-]([0-3]{1}[0-9]{1}).*?', date_)[0])

    return send_date_str


# 收取某封邮件
@retry(wait=wait_fixed(5), stop=stop_after_attempt(3))
def retrive(pop_conn, i):
    resp_s, lines, octets = pop_conn.retr(i)
    try:
        msg_content = b'\r\n'.join(lines).decode('utf-8')
        msg = Parser().parsestr(msg_content)
    except:
        msg = decode_email_content(lines)
    # 主题有效性验证
    if msg.get('Subject'):
        subject = decode_str(msg.get('Subject'))[0]
    else:
        subject = "主题不存在..."
    # 解析发件人邮箱地址
    hdr, addr = parseaddr(msg.get('From'))
    address = u'%s' % (addr)  # 发件人邮箱
    # 解析邮件日期
    date_ = msg.get('Date')
    send_date_str = validate_date(date_)
    return msg, address, subject, send_date_str


# 二分法获取某日期邮件的序号
def binary_search(mail_count_num, request_date):
    global pop_conn
    print('正在进入历史邮件快速查找,请稍等...')
    start = 1
    n = 0
    end = mail_count_num
    retrived_date = ''
    while start <= end:
        print('\n%s' % n)
        n = n + 1
        mid = (start + end) // 2
        print('Mid:%s' % mid)
        retrived_date = retrive(pop_conn, mid)[3]
        print('收取到的日期为:%s' % retrived_date)
        if request_date == retrived_date:
            return mid, n, retrived_date
        elif retrived_date > request_date:
            end = mid - 1
        else:
            start = mid + 1
    print('\n在第%s次查找后确认邮件位置...\n' % n)
    return start, n, retrived_date


# 精确查找表格中匹配目标字符串的第一个单元格,并返回坐标
def find_cell_value_r_c(open_sheet, strs):
    sheet = open_sheet
    rows = sheet.nrows  # 行数
    cols = sheet.ncols  # 列数
    # 每个要查找值的行数与列数
    for r in range(0, rows):
        for c in range(0, cols):
            cell_value = sheet.cell(r, c).value
            if cell_value == '':
                continue
            if strs == str(cell_value):
                return r, c


# 模糊查找表格中匹配目标字符串的第一个单元格,并返回坐标
def lazy_find_cell_value_r_c(open_sheet, strs):
    sheet = open_sheet
    rows = sheet.nrows  # 行数
    cols = sheet.ncols  # 列数
    # 每个要查找值的行数与列数
    for r in range(0, rows):
        for c in range(0, cols):
            cell_value = sheet.cell(r, c).value
            if cell_value == '':
                continue
            if strs in str(cell_value):
                return r, c


# 在第一列遍历, 精确查找到科目代码对应的行数

def find_cell_r(open_sheet, ke_mu_code):
    sheet = open_sheet
    rows = sheet.nrows
    for r in range(0, rows):
        cell_value = sheet.cell(r, 0).value
        if cell_value == '':
            continue
        if ke_mu_code == str(cell_value):
            return r


# 在第一列遍历, 模糊查找到科目代码对应的行数

def lazy_find_cell_r(open_sheet, excel_str_date):
    sheet = open_sheet
    rows = sheet.nrows
    for r in range(0, rows):
        cell_value = sheet.cell(r, 0).value
        if cell_value == '':
            continue
        if excel_str_date in str(cell_value):
            return r


# 下载附件, 解析出日期, 解析出所需的关键单元格
def download_attachment_parse(msg):
    global sheet
    for part in msg.walk():
        filename = part.get_filename()
        if filename:
            file_name_, char_ = decode_str(filename)
            if re.split('\.', file_name_)[len(re.split('\.', file_name_)) - 1] not in 'xlsx':
                print('【非excel附件】附件名:%s, 跳过...' % file_name_)
            else:
                if part.get_payload(decode=True):
                    data = part.get_payload(decode=True)
                    write_file = open(file_name_, 'wb')
                    write_file.write(data)
                    write_file.close()
                    print('附件%s已下载...' % file_name_)
                    # if re.split('\.', file_name_)[len(re.split('\.', file_name_))-1] not in 'xlsx':
                    #     print('不是excel类型...')
                    #     abs_file_path = os.path.join(os.getcwd(), file_name_)
                    #     os.remove(abs_file_path)  # 非excel附件, 删除
                    #     print('【已删除】附件%s非excel附件, 已删除...' % file_name_)
                    # else:

                    try:
                        open_workbook = xlrd.open_workbook(file_name_)  # 打开下载下来的excel
                    except:
                        pass
                    else:
                        open_sheet = open_workbook.sheet_by_index(0)  # 激活第一个sheet

                        target_values = []
                        # 获取银行存款
                        if find_cell_r(open_sheet, '1002') and find_cell_value_r_c(open_sheet, '市值'):
                            strs_row_bank = find_cell_r(open_sheet, '1002')  # 科目代码为'1002'银行存款的行数
                            strs_col = find_cell_value_r_c(open_sheet, '市值')[1]  # 市值的列数
                            strs_bank_value = open_sheet.cell(strs_row_bank,
                                                              strs_col).value  # 银行存款市值的单元格值****************************************
                        else:
                            strs_bank_value = '未在表格中找到[银行存款市值]'
                        # 获取单位净值
                        if lazy_find_cell_value_r_c(open_sheet, '单位净值') and lazy_find_cell_value_r_c(open_sheet, '单位净值'):
                            strs_net_r = lazy_find_cell_value_r_c(open_sheet, '单位净值')[0]  # 模糊查找单位净值的行数
                            strs_net_c = lazy_find_cell_value_r_c(open_sheet, '单位净值')[1]  # 模糊查找单位净值的列数
                            strs_net_value = open_sheet.cell(strs_net_r, strs_net_c).value  # 单位净值的单元格的值
                        else:
                            strs_net_value = '未在表格中找到[单位净值]'
                        # 获取估值日期
                        if lazy_find_cell_r(open_sheet, '日期'):

                            strs_date_r = lazy_find_cell_r(open_sheet, '日期')  # 模糊第一列查找日期的行数
                            strs_date_value = open_sheet.cell(strs_date_r, 0).value  # 获取日期估值日期
                        else:
                            strs_date_value = '未在表格中找到[估值日期]'

                        # 获取存出保证金
                        if find_cell_r(open_sheet, '1031') and find_cell_value_r_c(open_sheet, '市值'):  # 科目代码为'1031'银行存款的行数
                            strs_row_secure_cash = find_cell_r(open_sheet, '1031')  # 科目代码为'1031'银行存款的行数
                            strs_col = find_cell_value_r_c(open_sheet, '市值')[1]  # 市值的列数
                            strs_secure_cash_value = open_sheet.cell(strs_row_secure_cash,
                                                                     strs_col).value  # 存出保证金市值的单元格值***********************
                        else:
                            strs_secure_cash_value = '未在表格中找到[存出保证金]'

                        # 获取产品名称
                        product_name = file_name_

                        # 开始添加单元格值
                        # target_values.append(strs_date_value) # 估值日期
                        target_values.append(product_name)  # 产品名称
                        target_values.append(str(strs_bank_value))  # 银行存款
                        target_values.append(str(strs_secure_cash_value))  # 存出保证金
                        target_values.append(strs_net_value)  # 单位净值

                        # 第一列单元格内的日期的提取与标准化
                        pattern_1 = '.*?([1,2]{1}[0,9]{1}[0-9]{2})[/-]([0,1]{1}[0-9]{1})[/-]([0-3]{1}[0-9]{1}).*?'
                        pattern_2 = '.*?([1,2]{1}[0,9]{1}[0-9]{2}[0,1]{1}[0-9]{1}[0-3]{1}[0-9]{1}).*?'
                        re_strs_date_value = ''
                        if len(re.findall(pattern_1, strs_date_value)) != 0:
                            re_strs_date_value = ''.join(re.findall(pattern_1, strs_date_value)[0])
                        else:
                            if len(re.findall(pattern_2, strs_date_value)) != 0:
                                re_strs_date_value = re.findall(pattern_2, strs_date_value)[0]
                        # 释放xlrd资源, 关闭xlrd打开的表格
                        # open_workbook.release_resources()
                        # del open_workbook
                        abs_excel_file_path = os.path.join(os.getcwd(), file_name_)
                        if re_strs_date_value != request_date:
                            os.remove(abs_excel_file_path)  #
                            print('【已删除】附件%s估值日期[%s]不满足所需日期,...' % (file_name_, re_strs_date_value))
                        # 组装目标值到汇总excel表格里头
                        else:
                            print('【已保存】附件%s' % file_name_)
                            target_values.append(re_strs_date_value)
                            sheet.append(target_values)


# 主函数
if __name__ == '__main__':
    # 保存密码, 输入登陆信息
    if os.path.exists('user_info.txt'):
        with open('user_info.txt', 'r') as read_file:
            user_info = read_file.readlines()
            host = user_info[0].strip()
            user = user_info[1].strip()
            passwd = user_info[2].strip()
    else:
        host = input('请输入邮箱服务器地址:')
        user = input('请输入邮箱地址:')
        passwd = input('请输入邮箱密码:')
        with open('user_info.txt', 'w') as write_file:
            write_file.write(host + '\n')
            write_file.write(user + '\n')
            write_file.write(passwd + '\n')

    # 登录邮箱,返回对登录邮箱的引用
    mail_count_num, pop_conn = log_in(host, user, passwd)

    today = time.localtime(time.time())
    today_str = time.strftime("%Y%m%d", today)  # 今天日期格式化

    # 输入要处理邮件的标题日期
    request_date = input('请输入日期,收取今天邮件直接按回车(日期输入格式为20141213) :')  # 输入日期
    start = time.clock()
    if request_date == '':
        request_date = today_str

    currentpath = os.getcwd()
    foldername = request_date
    new_path = os.path.join(currentpath, foldername)  # 文件存储路径,字符串
    if os.path.exists(new_path) == False:  # 如果文件夹不存在，创建文件夹
        os.makedirs(new_path)
    os.chdir(new_path)

    wb = Workbook()
    sheet = wb.active
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 16
    sheet.column_dimensions['C'].width = 16
    sheet.column_dimensions['D'].width = 16
    sheet.column_dimensions['E'].width = 10
    table_head = ['产品名称', '银行存款', '存出保证金', '单位净值', '估值日期']
    sheet.append(table_head)
    # first_email_date = retrive(pop_conn, 1)[3]
    # dates_count = (datetime.datetime.strptime(today_str, '%Y%m%d') - datetime.datetime.strptime(first_email_date, '%Y%m%d')).days + 1
    # avg_email_per_day = mail_count_num // dates_count
    # print('邮箱总邮件数%s封, 收件箱邮件时间跨度%s天, 每天平均收取%s封邮件'%(mail_count_num,dates_count,avg_email_per_day))
    # 开始遍历
    if request_date == '':
        request_date = today_str
    if (datetime.datetime.strptime(today_str, '%Y%m%d') - datetime.datetime.strptime(request_date,
                                                                                     '%Y%m%d')).days >= 30:
        retri_num, fzd, retrived_date = binary_search(mail_count_num, request_date)
        print('开始从第[%s]封收取, 该封邮件收件日期为[%s]...' % (retri_num, retrived_date))
        for i in range(retri_num, mail_count_num, 1):
            print(i)
            msg, address, subject, send_date = retrive(pop_conn, i)
            print('\n收件日期:[%s]' % send_date)
            try:
                print('邮件标题:%s' % subject.encode('gbk').decode('gbk'))
            except:
                print('无法在控制台显示邮件标题')
            if (datetime.datetime.strptime(send_date, '%Y%m%d') - datetime.datetime.strptime(request_date,
                                                                                             '%Y%m%d')).days >= 5:
                print('\n--------请求日期未来五天内未没有收到所需邮件--------\n')
                break
            else:
                download_attachment_parse(msg)

    else:
        retri_num, fzd, retrived_date = binary_search(mail_count_num, request_date)
        print('开始从第[%s]封收取, 该封邮件收件日期为[%s]...' % (retri_num, retrived_date))
        for i in range(retri_num, mail_count_num, 1):
            print(i)
            msg, address, subject, send_date = retrive(pop_conn, i)
            print('收件日期:[%s]' % send_date)
            try:
                print('邮件标题:%s' % subject.encode('gbk').decode('gbk'))
            except:
                print('无法在控制台显示邮件标题')
            download_attachment_parse(msg)
            # else:
            #     for i in range(mail_count_num, 0, -1):
            #         print(i)
            #         msg, address, subject, send_date = retrive(pop_conn, i)
            #         download_attachment_parse(subject,msg)

    print('开始创建汇总表...')
    wb.save('%s汇总表.xlsx' % request_date)
    # pop_conn.quit()
    # print('任务完成, 10秒后将关闭窗口...')
    # time.sleep(10)
    print('===============任务完成===============\n')
    elapsed_time = time.clock()-start
    # print(elapsed_time)