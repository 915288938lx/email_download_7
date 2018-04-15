import openpyxl
import xlrd
import re
worksheet =xlrd.open_workbook('2017-6-12刘祥-王帅华客户梳理.xlsx').sheet_by_index(0)
re_strs_date_value = ''
pattern_1 = '.*?([1,2]{1}[0,9]{1}[0-9]{2})[/-]([0,1]{1}[0-9]{1})[/-]([0-3]{1}[0-9]{1}).*?'
pattern_2 = '.*?([1,2]{1}[0,9]{1}[0-9]{2}[0,1]{1}[0-9]{1}[0-3]{1}[0-9]{1}).*?'
# strs_date_value = ''

def lazy_find_cell_value_r_c(open_sheet, strs):
    sheet = open_sheet
    rows = sheet.nrows  # 行数
    cols = sheet.ncols  # 列数
    # 每个要查找值的行数与列数
    for r in range(0, rows):
        for c in range(0, cols):
            vv = sheet.cell(r, c).value
            if vv == '' :
                continue

            return r, c

lazy_find_cell_value_r_c(worksheet,'日期')





#
# def lazy_find_cell_r(open_sheet, excel_str_date):
#     sheet = open_sheet
#     rows = sheet.nrows
#     for r in range(0, rows):
#         if excel_str_date in sheet.cell(r, 0).value:
#             return r
#
# strs_date_r = lazy_find_cell_r(worksheet, '日期')  # 模糊第一列查找日期的行数
# strs_date_value = worksheet.cell(strs_date_r, 0).value  #
#
# if len(re.findall(pattern_1,strs_date_value)) != 0:
#     re_strs_date_value = ''.join(re.findall(pattern_1,strs_date_value)[0])
# else:
#     if len(re.findall(pattern_2, strs_date_value)) != 0:
#         re_strs_date_value = re.findall(pattern_2, strs_date_value)[0]
# print(re_strs_date_value)