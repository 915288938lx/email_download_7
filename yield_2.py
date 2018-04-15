import xlrd
from openpyxl import Workbook

# 查找表格所需单元格数据,并返回
str_list = ['单位净值', '日期', '市价', '银行存款', '存出保证金']
ke_mu_code = ['1002','1031']


# specific_col_item = ['']
# # 在第一列遍历, 查找到科目代码对应的行数的迭代器
#
# def find_cell_c(open_sheet, specific_col_item):
#     sheet = open_sheet
#     rows = sheet.nrows
#     for c in range(0, rows):
#         for s_p_i in specific_col_item:
#             if sheet.cell(c, 0).value == s_p_i:
#                 yield c
#
#在第一列遍历, 查找到科目代码对应的行数的迭代器

def find_cell_r(open_sheet,ke_mu_code):
    sheet = open_sheet
    rows = sheet.nrows
    for r in range(0,rows):
        for ke_mu in ke_mu_code:
            if sheet.cell(r,0).value == ke_mu:
                yield r


#获取第一列对应科目对应的行数索引
def get_row_index(it_r):
    li = []
    for r in it_r:
        li.append(r)
    print(li)
    return li


#返回满足传入列表的单元格的行和列坐标的迭代器, 遍历所有活动单元格
def find_cell_value_r_c(open_sheet, str_list):
    sheet = open_sheet
    rows = sheet.nrows  # 行数
    cols = sheet.ncols  # 列数
    # 每个要查找值的行数与列数
    for r in range(0, rows):
        for c in range(0, cols):
            for str_ in str_list:
                if str_ in sheet.cell(r, c).value:
                    yield r,c

#获取单元格的行和列的坐标
def get_row_col_index(it):
    li = []
    for x in it:
        li.append(x)
    print(li)
    return li






if __name__ == '__main__':
    li_openpy = []
    wb = Workbook()
    active_sheet = wb.active
    open_sheet = xlrd.open_workbook('11.xls').sheet_by_index(0)
    it = find_cell_value_r_c(open_sheet, str_list)
    it_r = find_cell_r(open_sheet, ke_mu_code)
    indx_r = get_row_index(it_r)
    indx = get_row_col_index(it)
    excel_date = open_sheet.cell(indx[0][0],indx[0][1]).value

    excel_net_value = open_sheet.cell(indx[1][0],indx[1][1]).value

    excel_cash = open_sheet.cell(indx[3][0],indx[2][1]+1).value

    excel_secure_cash = open_sheet.cell(indx[4][0],indx[2][1]+1).value

    li_openpy.append(excel_date)
    li_openpy.append(excel_net_value)
    li_openpy.append(excel_cash)
    li_openpy.append(excel_secure_cash)
    active_sheet.append(li_openpy)
    wb.save('huizongbiao.xlsx')
# # #
#
# it = find_cell_value_r_c(open_sheet, str_list)
# li = []
# for x in it:
#     li.append(x)
# print(li)
# print(li[0][0])
#     #
#     #
#     #
#     # li = []
#     # row_index = get_row_index(it)
#     # col_index = get_col_index(it)
#     # wb = Workbook('huizongbiao.xlsx')
#     # active_sheet = wb.active
#     # excel_date = it[0,1]
#     #
#     # excel_net_value =
#     #
#     # excel_cash_net =
#     #
#     # excel_secure_cash =
