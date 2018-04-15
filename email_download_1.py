import poplib
import time
import os
from email.parser import Parser
from email.header import decode_header
from email.utils import parseaddr
import xlrd
from openpyxl import Workbook
import re
import datetime

# 解码,解析加密传输的字符串成正常字符串
def decode_str(s):
    try:
        value, charset = decode_header(s)[0]
        if charset:
            value = value.decode(charset)
        return value
    except:
        pass


# 登录邮箱, 获取邮件字符串源文件,解析出主题/发件人邮箱,发送时间, 是否带有附件
def log_in(host, user, passwd):
    # 开始登录
    pop_conn = poplib.POP3_SSL(host)
    pop_conn.user(user)
    pop_conn.pass_(passwd)
    mail_count = len(pop_conn.list()[1])
    # 显示邮箱状态：邮件数量，占用空间
    # print('Messages: %s. Size: %s' % pop_conn.stat())
    print('测试信息...\n邮箱登陆成功! \n收件箱共有%s封邮件,占用空间%s字节\n' % (mail_count, pop_conn.stat()[1]))
    return mail_count, pop_conn


# 收取某封邮件
def retrive(pop_conn, i):
    resp_s, lines, octets = pop_conn.retr(i)
    # 解码原始邮件
    try:
        msg_content = b'\r\n'.join(lines).decode('utf-8')
        msg = Parser().parsestr(msg_content)

    except:
        li = []
        for line in lines:
            try:
                line_str = line.decode('utf-8')
                li.append(line_str)
            except:
                pass
        li_ = '\r\n'.join(li)
        msg = Parser().parsestr(li_)

    subject_ = msg.get('Subject')
    subject = decode_str(subject_)
    sender = msg.get('From')
    date1 = time.strptime(msg.get("Date")[0:24], '%a, %d %b %Y %H:%M:%S')  # 格式化收件时间
    send_date_str = time.strftime("%Y%m%d", date1)
    hdr, addr = parseaddr(sender)
    address = u'%s' % (addr)  # 发件人邮箱
    print(subject)
    return msg, address, subject, send_date_str
    pass




# 二分法获取某日期邮件的序号
def binary_search(mail_count_num, request_date):
    start = 0
    n = 0
    end = mail_count_num
    while start <= end:
        n = n+1
        mid = (start + end) // 2
        send_date =retrive(pop_conn,mid)[3]
        if send_date == request_date:
            return mid, send_date
        elif retrive(pop_conn, mid)[3] > request_date:
            end = mid - 1
        else:
            start = mid + 1
    print('执行了%s'%n)
    return start, retrive(pop_conn,start)[3]


# 获取标题中的发件日期
def get_title_date(address, subject):
    if address in ['zctgsjfs@tg.gtja.com', '915288938@qq.com']:
        title_date_ = re.findall(pattern_1, subject)
        return title_date_
    elif address in 'duanjiushuang@chinastock.com.cn':
        title_date_ = ''.join(re.findall(pattern_2, subject))
        return title_date_
    elif address in ['yangfan1@swhysc.com']:
        title_date_ = ''.join(re.findall(pattern_3, subject))
        return title_date_
    elif address in 'zszqwbfw@stocke.com.cn':
        title_date_ = ''.join(re.findall(pattern_4, subject))
        return title_date_
    elif address in 'dfjjwb@orientsec.com.cn':
        title_date_ = re.findall(pattern_5, subject)
        return title_date_
    elif address in 'cpbbfs@gfund.com':
        title_date_ = re.findall(pattern_6, subject)
        return title_date_
    elif address in 'FAreport@citics.com':
        title_date_ = re.findall(pattern_7, subject)
        return title_date_
    elif address in ['yywbfa@cmschina.com.cn']:
        title_date_ = re.findall(pattern_8, subject)
        return title_date_


# 检查该邮件是否满足要求日期,并下载附件到目录,同时读取该附件获得该表格对象
def down_parse_attachment(address, msg, subject):
    # 下载附件
    for part in msg.walk():
        filename = part.get_filename()
        if filename:
            file_name_ = decode_str(filename)
            data = part.get_payload(decode=True)
            write_file = open(file_name_, 'wb')
            write_file.write(data)
            write_file.close()
            print('邮件主题为:%s' % subject)
            print('附件%s已下载...' % file_name_)
            print('附件下载成功')
            # 读取附件表格
            open_workbook = xlrd.open_workbook(file_name_)
            print('开始解析保存下来的excel文件')
            open_sheet = open_workbook.sheet_by_index(0)
            lsheet = open_sheet
            # 国泰君安
            if address in ['zctgsjfs@tg.gtja.com', '915288938@qq.com']:
                try:
                    parse_gtja(lsheet)
                except:
                    pass
            # 银河证券
            elif address in ['duanjiushuang@chinastock.com.cn']:
                try:
                    parse_yhzq(lsheet)
                except:
                    pass
            # 申万,国金,东方,浙商
            elif address in ['zszqwbfw@stocke.com.cn', 'dfjjwb@orientsec.com.cn', 'cpbbfs@gfund.com',
                             'yangfan1@swhysc.com']:
                try:
                    parse_swhy(lsheet)
                except:
                    pass
            # 招商证券
            elif address in ['yywbfa@cmschina.com.cn']:
                try:
                    parse_zszq_(lsheet)
                except:
                    pass
            #
            elif address in ['FAreport@citics.com']:
                try:
                    parse_zxzq(lsheet)
                except:
                    pass


# 各个券商附件产品名字的正则表达式

# 国泰君安产品名称正则匹配表达式
reg_1 = re.compile('.*?___(.*?)___.*?', re.S)
# 银河证券产品名称正则匹配表达式
reg_2 = re.compile('.*?__.*?__(.*?)__.*?', re.S)
# 申万宏源证券产品名称正则表达式
reg_3 = re.compile('.*?___(.*?)___.*?', re.S)
# 浙商证券产品名称正则表达式
reg_4 = re.compile('.*?___(.*?)___.*?', re.S)
# 东方证券
reg_5 = re.compile('.*?___(.*?)___.*?', re.S)
# 国金证券产品名称正则表达式
reg_6 = re.compile('.*?___(.*?)___.*?', re.S)
# 中信证券产品名称正则表达式
reg_7 = re.compile('.*?___(.*?)___.*?', re.S)
# 招商证券产品名称正则表达式
reg_8 = re.compile('.*?__(.*?)__.*?', re.S)

# 各个邮件主题日期正则表达式
# 国泰君安
pattern_1 = re.compile('.*?私募证券投资基金([0-9]{8})', re.S)
# 银河证券
pattern_2 = re.compile('.*?私募证券投资基金([0-9]{4})年([0-9]{2})月([0-9]{2})日.*?', re.S)
# 申万宏源
pattern_3 = re.compile('.*?私募证券投资基金_([0-9]{4})-([0-9]{2})-([0-9]{2})', re.S)
# 浙商证券
pattern_4 = re.compile('.*?私募证券投资基金_([0-9]{4})-([0-9]{2})-([0-9]{2}).*?', re.S)
# 东方证券
pattern_5 = re.compile('([0-9]{8}).*?', re.S)
# 国金证券
pattern_6 = re.compile('.*?私募证券投资基金_([0-9]{8})', re.S)
# 中信证券
pattern_7 = re.compile('.*?估值表_([0-9]{8})', re.S)
# 招商证券
pattern_8 = re.compile('.*?私募证券投资基金_([0-9]{8})', re.S)


# 解析国泰君安表格数据
def parse_gtja(lsheet):
    print('正在解析国泰君安表格 ')
    li = []
    name = lsheet.cell(1, 0).value
    name_ = str(re.findall(reg_1, name)[0])
    valued_date = lsheet.cell(2, 0).value[-8:]
    bank = lsheet.cell(4, 7).value
    if lsheet.cell(7, 1).value != '存出保证金':
        cash = lsheet.cell(10, 7).value
    else:
        cash = lsheet.cell(7, 7).value
    net_value = ''.join(str(lsheet.cell(2, 7).value).split('单位净值：'))
    li.append(name_)
    li.append(valued_date)
    li.append(net_value)
    li.append(bank)
    li.append(cash)

    sheet.append(li)


# 读取申万/浙商/东方表格数据
def parse_swhy(lsheet):
    print('正在解析申万宏源/国金/东方证券/浙商表格 ')
    li = []
    name = lsheet.cell(1, 0).value
    print('name为...')
    print(name)
    name_ = str(re.findall(reg_3, name)[0])
    valued_date = ''.join(lsheet.cell(2, 0).value[-10:].split('-'))
    bank = str(lsheet.cell(4, 7).value)
    cash = str(lsheet.cell(7, 7).value)
    net_value = ''.join(str(lsheet.cell(2, 7).value).split('单位净值：'))
    li.append(name_)
    li.append(valued_date)
    li.append(net_value)
    li.append(bank)
    li.append(cash)
    print('读取中')
    sheet.append(li)
    print('正在写入汇总数据')


# 读取银河证券数据
def parse_yhzq(lsheet):
    print('正在解析银河证券表格 ')
    li = []
    name = lsheet.cell(2, 0).value
    name_ = str(re.findall(reg_2, name)[0])
    valued_date = ''.join(lsheet.cell(3, 0).value[-10:].split('-'))
    bank = lsheet.cell(10, 7).value
    cash = lsheet.cell(9, 7).value
    net_value = ''.join(str(lsheet.cell(3, 13).value).split('单位净值：'))
    li.append(name_)
    li.append(valued_date)
    li.append(net_value)
    li.append(bank)
    li.append(cash)

    sheet.append(li)


# 解析招商证券表格数据
def parse_zszq_(lsheet):
    print('正在解析招商证券表格 ')
    li = []
    name = lsheet.cell(2, 0).value
    name_ = str(re.findall(reg_8, name)[0])
    valued_date = ''.join(lsheet.cell(3, 0).value[-10:].split('-'))
    bank = str(lsheet.cell(8, 9).value)
    cash = str(lsheet.cell(11, 9).value)
    net_value = ''.join(str(lsheet.cell(3, 10).value).split('单位净值：'))
    li.append(name_)
    li.append(valued_date)
    li.append(net_value)
    li.append(bank)
    li.append(cash)

    sheet.append(li)


# 解析中信证券表格数据
def parse_zxzq(lsheet):
    print('正在解析中信证券表格 ')
    li = []
    name = lsheet.cell(2, 0).value
    name_ = str(re.findall(reg_7, name)[0])
    valued_date = ''.join(lsheet.cell(3, 0).value[-10:].split('-'))
    bank = str(lsheet.cell(7, 11).value)
    cash = str(lsheet.cell(10, 11).value)
    net_value = ''.join(str(lsheet.cell(3, 10).value).split('单位净值:'))
    li.append(name_)
    li.append(valued_date)
    li.append(net_value)
    li.append(bank)
    li.append(cash)

    sheet.append(li)


# 主函数
if __name__ == '__main__':
    # 保存密码, 输入登陆信息
    if os.path.exists('user_info.txt'):
        with open('user_info.txt', 'r') as read_file:
            user_info = read_file.readlines()
            host = user_info[0].strip()
            user = user_info[1].strip()
            passwd = user_info[2].strip()
    else:
        host = input('请输入邮箱服务器地址:')
        user = input('请输入邮箱地址:')
        passwd = input('请输入邮箱密码:')
        with open('user_info.txt', 'w') as write_file:
            write_file.write(host + '\n')
            write_file.write(user + '\n')
            write_file.write(passwd + '\n')

    #登录邮箱
    mail_count_num, pop_conn = log_in(host, user, passwd)

    # 输入要处理邮件的标题日期
    request_date = input('请输入日期,收取今天邮件直接按回车(日期输入格式为20141213) :')  # 输入日期
    today = time.localtime(time.time())
    today_str = time.strftime("%Y%m%d", today)  # 今天日期格式化


    # 根据输入的日期, 建立目录,创建汇总表
    currentpath = os.getcwd()  # 获取当前目录
    foldername = request_date  # 文件夹名和最后输出文件名
    new_path = os.path.join(currentpath, foldername)  # 文件存储路径,字符串
    if os.path.exists(new_path) == False:  # 如果文件夹不存在，创建文件夹
        os.makedirs(new_path)
    # 切换工作目录
    os.chdir(new_path)
    wb = Workbook()
    sheet = wb.active
    sheet.column_dimensions['A'].width = 42
    sheet.column_dimensions['B'].width = 9.2
    sheet.column_dimensions['C'].width = 8.4
    sheet.column_dimensions['D'].width = 12
    sheet.column_dimensions['E'].width = 12
    table_head = ['产品名称', '估值日期', '单位净值', '银行存款', '证券余额']
    sheet.append(table_head)

    #开始遍历
    skipped_num = 0
    if request_date == '':
        request_date = today_str
    if (datetime.datetime.strptime(today_str,'%Y%m%d') - datetime.datetime.strptime(request_date,'%Y%m%d')).days >= 7:
        # print('开始二分法快速查找指定日期...')
        retri_num , send_date= binary_search(mail_count_num, request_date)[0]
        print('开始从第')
        for i in range(retri_num, mail_count_num, 1):
            msg, address, subject, send_date = retrive(log_in(host, user, passwd)[1], i)
            if get_title_date(address, subject):
                title_date = get_title_date(address, subject)[0]
                if title_date == request_date:
                    down_parse_attachment(address, msg, request_date)
                else:
                    skipped_num = skipped_num+1
            else:
                skipped_num = skipped_num + 1
    else:
        for i in range(mail_count_num, 0, -1):
            msg, address, subject, send_date = retrive(pop_conn, i)
            if get_title_date(address, subject):
                title_date = get_title_date(address, subject)[0]
                if title_date == request_date:
                    down_parse_attachment(address, msg, request_date)
                else:
                    skipped_num = skipped_num + 1
            else:
                skipped_num = skipped_num + 1



    wb.save('%s汇总表.xlsx' % foldername)
    print('===============任务完成===============\n')
