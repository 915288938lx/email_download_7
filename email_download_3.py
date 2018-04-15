import poplib
poplib._MAXLINE = 20480
import time
import os
from email.parser import Parser
from email.header import decode_header
from email.utils import parseaddr
from openpyxl import Workbook
import re
import datetime
from tenacity import retry, wait_fixed, stop_after_attempt
import xlrd



# 查找表格所需单元格数据,并返回
def find_cell_value_r_c(open_sheet, str_list):
    sheet = open_sheet
    rows = sheet.nrows  # 行数
    cols = sheet.ncols  # 列数
    #每个要查找值的行数与列数
    row_col = []
    #将row_col的列表再装配到row_col_list中
    row_col_list = []
    for r in range(0, rows):
        for c in range(0, cols):
            got_cell_value = sheet.cell(r, c).value
            for str_ in str_list:
                if str_ in got_cell_value:
                    str_row = r
                    str_col = c
                    row_col.append(str_row)
                    row_col.append(str_col)
                    row_col_list.append(row_col)
                else:
                    print('未找到要查找的值...')
    return row_col_list

# 解码,解析加密传输的字符串成正常字符串
def decode_str(s):
    try:
        value, charset = decode_header(s)[0]
        if charset:
            value = value.decode(charset)
        return value
    except:
        pass

@retry(wait=wait_fixed(3), stop=stop_after_attempt(3))
# 登录邮箱, 获取邮件字符串源文件,解析出主题/发件人邮箱,发送时间, 是否带有附件
def log_in(host, user, passwd):
    # 开始登录
    pop_conn = poplib.POP3_SSL(host)
    pop_conn.user(user)
    pop_conn.pass_(passwd)
    pop_conn.noop()
    mail_count = len(pop_conn.list()[1])
    # 显示邮箱状态：邮件数量，占用空间
    # print('Messages: %s. Size: %s' % pop_conn.stat())
    # print('测试信息...\n邮箱登陆成功! \n收件箱共有%s封邮件,占用空间%s字节\n' % (mail_count, pop_conn.stat()[1]))
    return mail_count, pop_conn


# 收取某封邮件
def retrive(pop_conn, i):
    try:
        resp_s, lines, octets = pop_conn.retr(i)
    except:
        resp_s, lines, octets = log_in(host, user, passwd)[1].retr(i)
    # 解码原始邮件
    try:
        msg_content = b'\r\n'.join(lines).decode('utf-8')
        msg = Parser().parsestr(msg_content)

    except:
        li = []
        for line in lines:
            try:
                line_str = line.decode('utf-8')
                li.append(line_str)
            except:
                pass
        li_ = '\r\n'.join(li)
        msg = Parser().parsestr(li_)
    # 主题有效性验证
    if msg.get('Subject'):
        subject = decode_str(msg.get('Subject'))
    else:
        subject = "None"
    # 日期有效性验证
    try:
        time.strptime(msg.get("Date")[0:24], '%a, %d %b %Y %H:%M:%S')
        date1 = time.strptime(msg.get("Date")[0:24], '%a, %d %b %Y %H:%M:%S')  # 格式化收件时间
        send_date_str = time.strftime("%Y%m%d", date1)
    except:
        send_date_str = "None"
        pass

    # 发件人有效性验证
    if msg.get('From'):
        hdr, addr = parseaddr(msg.get('From'))
        address = u'%s' % (addr)  # 发件人邮箱
    else:
        address = "2"

    return msg, address, subject, send_date_str


# 二分法获取某日期邮件的序号
def binary_search(mail_count_num, request_date):
    global pop_conn
    print('正在进行二分快速查找,请稍后...')
    start = 0
    n = 0
    end = mail_count_num
    while start <= end:
        n = n + 1
        mid = (start + end) // 2
        if retrive(pop_conn, mid)[3] == "None":
            continue
        send_date = retrive(pop_conn, mid)[3]
        if send_date == request_date:
            return mid, n
        elif retrive(pop_conn, mid)[3] > request_date:
            end = mid - 1
        else:
            start = mid + 1
    print('二分法时间复杂度为%s次' % n)
    return start, n


# 获取标题中的发件日期
def get_title_date(address, subject):
    # 国泰君安标题日期获取
    if address in ['zctgsjfs@tg.gtja.com']:
        title_date_ = re.findall(pattern_1, subject)
        return title_date_
    # 银河证券标题日期获取
    elif address in 'duanjiushuang@chinastock.com.cn':
        title_date_ = ''.join(re.findall(pattern_2, subject))
        return title_date_
    # 申万宏源标题日期获取
    elif address in ['yangfan1@swhysc.com']:
        title_date_ = ''.join(re.findall(pattern_3, subject))
        return title_date_
    # 浙商证券标题日期获取
    elif address in 'zszqwbfw@stocke.com.cn':
        title_date_ = ''.join(re.findall(pattern_4, subject))
        return title_date_
    # 东方证券标题日期获取
    elif address in 'dfjjwb@orientsec.com.cn':
        title_date_ = re.findall(pattern_5, subject)
        return title_date_
    # 国金证券标题日期获取
    elif address in 'cpbbfs@gfund.com':
        title_date_ = re.findall(pattern_6, subject)
        return title_date_
    # 中信证券标题日期获取
    elif address in 'FAreport@citics.com':
        title_date_ = re.findall(pattern_7, subject)
        return title_date_
    # 招商证券标题日期获取
    elif address in ['yywbfa@cmschina.com.cn']:
        title_date_ = re.findall(pattern_8, subject)
        return title_date_
    # 兴业期货标题日期获取
    elif address in ['js@cifutures.com.cn']:
        title_date_ = re.findall(pattern_9, subject)
        return title_date_
    # 中金期货标题日期获取
    elif address in ['cpbbfs@gfund.com']:
        title_date_ = re.findall(pattern_9, subject)
        return title_date_
    # 平安资管标题日期获取
    elif address in ['admin@service.pingan.com']:
        title_date_ = re.findall(pattern_10, subject)
        return title_date_
    # 华润信托标题日期获取
    elif address in ['crtliangy@crctrust.com']:
        title_date_ = re.findall(pattern_11, subject)
        return title_date_
    # 兴业期货标题日期获取
    elif address in ['yezs@cifutures.com.cn']:
        title_date_ = re.findall(pattern_12, subject)
        return title_date_
    # 华泰期货标题日期获取
    elif address in ['zggzhd@htfc.com']:
        title_date_ = re.findall(pattern_13, subject)
        return title_date_


# 检查该邮件是否满足要求日期,并下载附件到目录,同时读取该附件获得该表格对象
def down_parse_attachment(msg, subject):
    global sheet
    # 下载附件
    for part in msg.walk():
        filename = part.get_filename()
        if filename:
            file_name_ = decode_str(filename)
            data = part.get_payload(decode=True)
            write_file = open(file_name_, 'wb')
            write_file.write(data)
            write_file.close()
            print('邮件主题为:%s' % subject)
            print('附件%s已下载...' % file_name_)
            print('附件下载成功')
            # 读取附件表格
            append_rows = []
            open_workbook = xlrd.open_workbook(file_name_)
            print('开始解析附件...')
            open_sheet = open_workbook.sheet_by_index(0)
            excel_date=open_sheet.cell(find_cell_value_r_c(open_sheet,str_list)[0][0],find_cell_value_r_c(open_sheet,str_list)[0][1]).value
            excel_net_value = open_sheet.cell(find_cell_value_r_c(open_sheet,str_list)[1][0],find_cell_value_r_c(open_sheet,str_list)[1][1]).value
            excel_cash_net = open_sheet.cell(find_cell_value_r_c(open_sheet,str_list)[3][0],find_cell_value_r_c(open_sheet,str_list)[2][1]).value
            excel_secure_cash = open_sheet.cell(find_cell_value_r_c(open_sheet,str_list)[4][0],find_cell_value_r_c(open_sheet,str_list)[2][1]).value
            append_rows.append(excel_date)
            append_rows.append(excel_net_value)
            append_rows.append(excel_cash_net)
            append_rows.append(excel_secure_cash)
            sheet.append(append_rows)

# 各个邮件主题日期正则表达式

# 国泰君安
pattern_1 = re.compile('.*?私募证券投资基金([0-9]{8})', re.S)
# 银河证券
pattern_2 = re.compile('.*?私募证券投资基金([0-9]{4})年([0-9]{2})月([0-9]{2})日.*?', re.S)
# 申万宏源
pattern_3 = re.compile('.*?私募证券投资基金_([0-9]{4})-([0-9]{2})-([0-9]{2})', re.S)
# 浙商证券
pattern_4 = re.compile('.*?私募证券投资基金_([0-9]{4})-([0-9]{2})-([0-9]{2}).*?', re.S)
# 东方证券
pattern_5 = re.compile('([0-9]{8}).*?', re.S)
# 国金证券
pattern_6 = re.compile('.*?私募证券投资基金_([0-9]{8})', re.S)
# 中信证券
pattern_7 = re.compile('.*?估值表_([0-9]{8})', re.S)
# 招商证券
pattern_8 = re.compile('.*?私募证券投资基金_([0-9]{8})', re.S)
# 兴业期货
pattern_9 = re.compile('.*?私募证券投资基金_([0-9]{8})', re.S)
# 中金期货
pattern_10 = re.compile('.*?_([0-9]{8})', re.S)
# 平安资管
pattern_11 = re.compile('.*?-([0-9]{4})-([0-9]{2})-([0-9]{2})', re.S)
# 华润信托
pattern_12 = re.compile('.*?估值表([0-9]{8})', re.S)
# 兴业期货
pattern_13 = re.compile('.*?_([0-9]{4})-([0-9]{2})-([0-9]{2}).*?', re.S)




# 主函数
if __name__ == '__main__':
    # 保存密码, 输入登陆信息
    if os.path.exists('user_info.txt'):
        with open('user_info.txt', 'r') as read_file:
            user_info = read_file.readlines()
            host = user_info[0].strip()
            user = user_info[1].strip()
            passwd = user_info[2].strip()
    else:
        host = input('请输入邮箱服务器地址:')
        user = input('请输入邮箱地址:')
        passwd = input('请输入邮箱密码:')
        with open('user_info.txt', 'w') as write_file:
            write_file.write(host + '\n')
            write_file.write(user + '\n')
            write_file.write(passwd + '\n')
    #要查找的值的列表
    str_list = ['日期','单位净值','市值','银行存款','存出保证金']

    # 登录邮箱,返回对登录邮箱的引用
    mail_count_num, pop_conn = log_in(host, user, passwd)

    # 输入要处理邮件的标题日期
    request_date = input('请输入日期,收取今天邮件直接按回车(日期输入格式为20141213) :')  # 输入日期
    today = time.localtime(time.time())
    today_str = time.strftime("%Y%m%d", today)  # 今天日期格式化

    # 根据输入的日期, 建立目录,创建汇总表
    currentpath = os.getcwd()  # 获取当前目录
    foldername = request_date  # 文件夹名和最后输出文件名
    new_path = os.path.join(currentpath, foldername)  # 文件存储路径,字符串
    if os.path.exists(new_path) == False:  # 如果文件夹不存在，创建文件夹
        os.makedirs(new_path)
    # 切换工作目录,创建汇总工作簿.xlsx
    os.chdir(new_path)
    wb = Workbook()
    sheet = wb.active
    sheet.column_dimensions['A'].width = 42
    sheet.column_dimensions['B'].width = 9.2
    sheet.column_dimensions['C'].width = 8.4
    sheet.column_dimensions['D'].width = 12
    sheet.column_dimensions['E'].width = 12
    table_head = ['产品名称', '估值日期', '单位净值', '银行存款', '证券余额']
    sheet.append(table_head)

    # 开始遍历
    skipped_num = 0
    if request_date == '':
        request_date = today_str
    if (datetime.datetime.strptime(today_str, '%Y%m%d') - datetime.datetime.strptime(request_date, '%Y%m%d')).days >= 7:
        retri_num, fzd = binary_search(mail_count_num, request_date)
        print('开始从第%s封收取...' % retri_num)
        for i in range(retri_num, mail_count_num, 1):
            msg, address, subject, send_date = retrive(pop_conn, i)
            print(subject)
            if get_title_date(address, subject):
                title_date = get_title_date(address, subject)[0]
                if title_date == request_date:
                    print('下载邮件...')
                    down_parse_attachment(msg, subject)
                    print('下载完毕...')
                else:
                    skipped_num = skipped_num + 1
                    print('标题日期不匹配,跳过...')
            else:
                skipped_num = skipped_num + 1
                print('标题日期不存在,跳过...')
    else:
        for i in range(mail_count_num, 0, -1):
            msg, address, subject, send_date = retrive(pop_conn, i)
            print(subject)
            if get_title_date(address, subject):
                title_date = get_title_date(address, subject)[0]
                if title_date == request_date:
                    down_parse_attachment(msg, subject)
                else:
                    skipped_num = skipped_num + 1
                    print('标题日期不匹配,跳过...')
            else:
                skipped_num = skipped_num + 1
                print('标题日期不存在,跳过...')
    print('一共跳过%s封邮件' % skipped_num)

    wb.save('%s汇总表.xlsx' % foldername)
    print('===============任务完成===============\n')
    pop_conn.quit()
