import poplib
poplib._MAXLINE = 20480
import time
import os
from email.parser import Parser
from email.header import decode_header
from email.utils import parseaddr, parsedate
from openpyxl import Workbook
import re
import datetime
from tenacity import retry, wait_fixed, stop_after_attempt
import xlrd

charset_collect = ['gb2312', 'gbk', 'gb18030', 'iso-8859-2', 'big5', 'iso-8859-6', 'iso-8859-4', 'iso-8859-5',
                   'iso-8859-7', 'iso-2022-kr', 'iso-8859-3', 'iso-2022-jp', 'iso-2022-jp', 'iso-8859-15', 'iso-8859-9',
                   'iso-8859-8-i', 'iso-8859-8', 'iso-8859-1', 'unicode', 'csISO2022JP', 'x-Chinese-CNS',
                   'x-Chinese-Eten', 'x-mac-chinesetrad', 'macintosh', 'ibm857', 'windows-1254', 'windows-1258',
                   'Windows-1252', 'windows-874', 'hz-gb-2312', 'windows-1256', 'ibm775', 'windows-1257', 'ibm852',
                   'x-mac-ce', 'windows-1250', 'EUC-CN', 'x-mac-chinesesimp', 'cp866', 'koi8-r', 'koi8-u',
                   'x-mac-cyrillic', 'windows-1251', 'x-Europa', 'x-IA5-German', 'ibm737', 'x-mac-greek',
                   'windows-1253', 'ibm869', 'DOS-862', 'x-mac-hebrew', 'windows-1255', 'ASMO-708', 'DOS-720',
                   'x-mac-arabic', 'x-EBCDIC-Arabic', 'x-EBCDIC-CyrillicRussian', 'x-EBCDIC-CyrillicSerbianBulgarian',
                   'x-EBCDIC-DenmarkNorway', 'x-ebcdic-denmarknorway-euro', 'x-EBCDIC-FinlandSweden',
                   'x-ebcdic-finlandsweden-euro', 'x-ebcdic-finlandsweden-euro', 'x-ebcdic-france-euro',
                   'x-EBCDIC-Germany', 'x-ebcdic-germany-euro', 'x-EBCDIC-GreekModern', 'x-EBCDIC-Greek',
                   'x-EBCDIC-Hebrew', 'x-EBCDIC-Icelandic', 'x-ebcdic-icelandic-euro', 'x-ebcdic-international-euro',
                   'x-EBCDIC-Italy', 'x-ebcdic-italy-euro', 'x-EBCDIC-JapaneseAndKana',
                   'x-EBCDIC-JapaneseAndJapaneseLatin', 'x-EBCDIC-JapaneseAndUSCanada', 'x-EBCDIC-JapaneseKatakana',
                   'x-EBCDIC-KoreanAndKoreanExtended', 'x-EBCDIC-KoreanExtended', 'CP870', 'x-EBCDIC-SimplifiedChinese',
                   'X-EBCDIC-Spain', 'x-ebcdic-spain-euro', 'x-EBCDIC-Thai', 'x-EBCDIC-TraditionalChinese', 'CP1026',
                   'x-EBCDIC-Turkish', 'x-EBCDIC-UK', 'x-ebcdic-uk-euro', 'ebcdic-cp-us', 'x-ebcdic-cp-us-euro',
                   'ibm861', 'x-mac-icelandic', 'x-iscii-as', 'x-iscii-be', 'x-iscii-de', 'x-iscii-gu', 'x-iscii-ka',
                   'x-iscii-ma', 'x-iscii-or', 'x-iscii-pa', 'x-iscii-ta', 'x-iscii-te', 'euc-jp', 'x-mac-japanese',
                   'shift_jis', 'ks_c_5601-1987', 'euc-kr', 'Johab', 'x-mac-korean', 'x-IA5-Norwegian', 'IBM437',
                   'x-IA5-Swedish', 'x-mac-turkish', 'unicodeFFFE', 'utf-7', 'us-ascii', 'ibm850', 'x-IA5']
address_list = ['tg.gtja.com', 'chinastock.com.cn', 'swhysc.com', 'stocke.com.cn', 'orientsec.com.cn', 'citics.com',
                'cmschina.com.cn', 'gfund.com', 'service.pingan.com', 'crctrust.com', 'cifutures.com.cn', 'htfc.com']

# 要查找的值的列表
str_list = ['日期', '单位净值', '市值', '银行存款', '存出保证金']

# 查找表格所需单元格数据,并返回
def find_cell_value_r_c(open_sheet, str_list):
    sheet = open_sheet
    rows = sheet.nrows  # 行数
    cols = sheet.ncols  # 列数
    # 每个要查找值的行数与列数
    for r in range(0, rows):
        for c in range(0, cols):
            for str_ in str_list:
                if str_ in sheet.cell(r, c).value:
                    yield r,c

# 解码,解析加密传输的字符串成正常字符串
def decode_str(s):
    try:
        value, charset = decode_header(s)[0]
        if charset:
            value = value.decode(charset)
        return value
    except:
        pass


@retry(wait=wait_fixed(3), stop=stop_after_attempt(3))
# 登录邮箱, 获取邮件字符串源文件,解析出主题/发件人邮箱,发送时间, 是否带有附件
def log_in(host, user, passwd):
    # 开始登录
    pop_conn = poplib.POP3_SSL(host)
    pop_conn.user(user)
    pop_conn.pass_(passwd)
    pop_conn.noop()
    mail_count = len(pop_conn.list()[1])
    # 显示邮箱状态：邮件数量，占用空间
    # print('Messages: %s. Size: %s' % pop_conn.stat())
    # print('测试信息...\n邮箱登陆成功! \n收件箱共有%s封邮件,占用空间%s字节\n' % (mail_count, pop_conn.stat()[1]))
    return mail_count, pop_conn


# 多字符集尝试解码邮件
def decode_email_content(lines):
    counter = 1
    str_lines = 0
    got_charset = ''
    for charset in charset_collect:
        try:
            str_lines = b'\r\n'.join(lines).decode(charset)
            got_charset = charset
            break
        except:
            counter = counter + 1
            continue
    print('在第%s次尝试后, 用%s字符集解码邮件成功' % (counter, got_charset))
    msg = Parser().parsestr(str_lines)
    return msg


# 收取某封邮件
@retry(wait=wait_fixed(5), stop=stop_after_attempt(3))
def retrive(pop_conn, i):
    resp_s, lines, octets = pop_conn.retr(i)
    try:
        msg_content = b'\r\n'.join(lines).decode('utf-8')
        msg = Parser().parsestr(msg_content)
    except:
        msg = decode_email_content(lines)
    # 主题有效性验证
    if msg.get('Subject'):
        subject = decode_str(msg.get('Subject'))
    else:
        subject = "主题不存在..."
    # 解析发件人邮箱地址
    hdr, addr = parseaddr(msg.get('From'))
    address = u'%s' % (addr)  # 发件人邮箱
    # 解析邮件日期
    date_ = msg.get('Date')
    if parsedate(date_):
        date_tuple = parsedate(date_)
        send_date_str = time.strftime("%Y%m%d", date_tuple)
    else:
        send_date_str = ''.join(
            re.findall('.*?([1,2]{1}[0,9]{1}[0-9]{2})[/-]([0,1]{1}[0-9]{1})[/-]([0-3]{1}[0-9]{1}).*?', date_)[0])

    return msg, address, subject, send_date_str


# 二分法获取某日期邮件的序号
def binary_search(mail_count_num, request_date):
    global pop_conn
    print('正在进入历史邮件快速查找,请稍等...')
    start = 0
    n = 0
    end = mail_count_num
    while start <= end:
        print('doing...')
        n = n + 1
        mid = (start + end) // 2
        print('mid%s' % mid)
        retrived_date = retrive(pop_conn, mid)[3]
        print('retrived_date为:%s' % retrived_date)
        if request_date == retrived_date:
            return mid, n
        elif retrived_date > request_date:
            end = mid - 1
        else:
            start = mid + 1
        print(n)
    print('二分法时间复杂度为%s次' % n)
    return start, n


# 获取标题中的发件日期
def get_title_date(address, subject):
    # 国泰君安标题日期获取
    if address in ['zctgsjfs@tg.gtja.com']:
        title_date_ = re.findall(pattern_1, subject)
        return title_date_
    # 银河证券标题日期获取
    elif address in ['duanjiushuang@chinastock.com.cn']:
        title_date_ = ''.join(re.findall(pattern_2, subject))
        return title_date_
    # 申万宏源标题日期获取
    elif address in ['yangfan1@swhysc.com']:
        title_date_ = ''.join(re.findall(pattern_3, subject))
        return title_date_
    # 浙商证券标题日期获取
    elif address in ['zszqwbfw@stocke.com.cn']:
        title_date_ = ''.join(re.findall(pattern_4, subject))
        return title_date_
    # 东方证券标题日期获取
    elif address in ['dfjjwb@orientsec.com.cn']:
        title_date_ = re.findall(pattern_5, subject)
        return title_date_
    # 国金证券标题日期获取
    elif address in ['cpbbfs@gfund.com']:
        title_date_ = re.findall(pattern_6, subject)
        return title_date_
    # 中信证券标题日期获取
    elif address in ['FAreport@citics.com']:
        title_date_ = re.findall(pattern_7, subject)
        return title_date_
    # 招商证券标题日期获取
    elif address in ['yywbfa@cmschina.com.cn']:
        title_date_ = re.findall(pattern_8, subject)
        return title_date_
    # 兴业期货标题日期获取
    elif address in ['js@cifutures.com.cn']:
        title_date_ = re.findall(pattern_9, subject)
        return title_date_
    # 中金期货标题日期获取
    elif address in ['cpbbfs@gfund.com']:
        title_date_ = re.findall(pattern_9, subject)
        return title_date_
    # 平安资管标题日期获取
    elif address in ['admin@service.pingan.com']:
        title_date_ = re.findall(pattern_10, subject)
        return title_date_
    # 华润信托标题日期获取
    elif address in ['crtliangy@crctrust.com']:
        title_date_ = re.findall(pattern_11, subject)
        return title_date_
    # 兴业期货标题日期获取
    elif address in ['yezs@cifutures.com.cn']:
        title_date_ = re.findall(pattern_12, subject)
        return title_date_
    # 华泰期货标题日期获取
    elif address in ['zggzhd@htfc.com']:
        title_date_ = re.findall(pattern_13, subject)
        return title_date_


# 检查该邮件是否满足要求日期,并下载附件到目录,同时读取该附件获得该表格对象
def download_attachment(msg):
    # 下载附件
    for part in msg.walk():
        filename = part.get_filename()
        if filename:
            file_name_ = decode_str(filename)
            data = part.get_payload(decode=True)
            write_file = open(file_name_, 'wb')
            write_file.write(data)
            write_file.close()
            print('附件%s已下载...' % file_name_)
            parse_excel(file_name_)
        else:
            pass

# 解析excel附件
def parse_excel(file_name_):
    # 读取附件表格
    global sheet
    append_rows = []
    open_workbook = xlrd.open_workbook(file_name_)
    print('开始解析附件...')
    open_sheet = open_workbook.sheet_by_index(0)
    it = find_cell_value_r_c(open_sheet, str_list)
    for x in it:
        excel_date =





    # excel_date = open_sheet.cell(find_cell_value_r_c(open_sheet, str_list)[0][0],
    #                              find_cell_value_r_c(open_sheet, str_list)[0][1]).value
    # excel_net_value = open_sheet.cell(find_cell_value_r_c(open_sheet, str_list)[1][0],
    #                                   find_cell_value_r_c(open_sheet, str_list)[1][1]).value
    # excel_cash_net = open_sheet.cell(find_cell_value_r_c(open_sheet, str_list)[3][0],
    #                                  find_cell_value_r_c(open_sheet, str_list)[2][1]).value
    # excel_secure_cash = open_sheet.cell(find_cell_value_r_c(open_sheet, str_list)[4][0],
    #                                     find_cell_value_r_c(open_sheet, str_list)[2][1]).value
    append_rows.append(excel_date)
    append_rows.append(excel_net_value)
    append_rows.append(excel_cash_net)
    append_rows.append(excel_secure_cash)
    sheet.append(append_rows)
    print('附件%s已解析' % file_name_)


# 各个邮件主题日期正则表达式

# 国泰君安
pattern_1 = re.compile('.*?私募证券投资基金([0-9]{8})', re.S)
# 银河证券
pattern_2 = re.compile('.*?私募证券投资基金([0-9]{4})年([0-9]{2})月([0-9]{2})日.*?', re.S)
# 申万宏源
pattern_3 = re.compile('.*?私募证券投资基金_([0-9]{4})-([0-9]{2})-([0-9]{2})', re.S)
# 浙商证券
pattern_4 = re.compile('.*?私募证券投资基金_([0-9]{4})-([0-9]{2})-([0-9]{2}).*?', re.S)
# 东方证券
pattern_5 = re.compile('([0-9]{8}).*?', re.S)
# 国金证券
pattern_6 = re.compile('.*?私募证券投资基金_([0-9]{8})', re.S)
# 中信证券
pattern_7 = re.compile('.*?估值表_([0-9]{8})', re.S)
# 招商证券
pattern_8 = re.compile('.*?私募证券投资基金_([0-9]{8})', re.S)
# 兴业期货
pattern_9 = re.compile('.*?私募证券投资基金_([0-9]{8})', re.S)
# 中金期货
pattern_10 = re.compile('.*?_([0-9]{8})', re.S)
# 平安资管
pattern_11 = re.compile('.*?-([0-9]{4})-([0-9]{2})-([0-9]{2})', re.S)
# 华润信托
pattern_12 = re.compile('.*?估值表([0-9]{8})', re.S)
# 兴业期货
pattern_13 = re.compile('.*?_([0-9]{4})-([0-9]{2})-([0-9]{2}).*?', re.S)

# 主函数
if __name__ == '__main__':
    # 保存密码, 输入登陆信息
    if os.path.exists('user_info.txt'):
        with open('user_info.txt', 'r') as read_file:
            user_info = read_file.readlines()
            host = user_info[0].strip()
            user = user_info[1].strip()
            passwd = user_info[2].strip()
    else:
        host = input('请输入邮箱服务器地址:')
        user = input('请输入邮箱地址:')
        passwd = input('请输入邮箱密码:')
        with open('user_info.txt', 'w') as write_file:
            write_file.write(host + '\n')
            write_file.write(user + '\n')
            write_file.write(passwd + '\n')


    # 登录邮箱,返回对登录邮箱的引用
    mail_count_num, pop_conn = log_in(host, user, passwd)

    # 输入要处理邮件的标题日期
    request_date = input('请输入日期,收取今天邮件直接按回车(日期输入格式为20141213) :')  # 输入日期
    today = time.localtime(time.time())
    today_str = time.strftime("%Y%m%d", today)  # 今天日期格式化

    # 根据输入的日期, 建立目录,创建汇总表
    currentpath = os.getcwd()  # 获取当前目录
    foldername = request_date  # 即将要创建的日期文件夹名
    new_path = os.path.join(currentpath, foldername)  # 在当前路径创建日期文件夹
    if os.path.exists(new_path) == False:  # 如果文件夹不存在，创建文件夹
        os.makedirs(new_path)
    # 切换到新创建的目录下,创建汇总工作簿.xlsx
    os.chdir(new_path)
    wb = Workbook()
    sheet = wb.active
    sheet.column_dimensions['A'].width = 42
    sheet.column_dimensions['B'].width = 9.2
    sheet.column_dimensions['C'].width = 8.4
    sheet.column_dimensions['D'].width = 12
    sheet.column_dimensions['E'].width = 12
    table_head = ['产品名称', '估值日期', '单位净值', '银行存款', '证券余额']
    sheet.append(table_head)

    # 开始遍历
    skipped_num = 0
    if request_date == '':
        request_date = today_str
    if (datetime.datetime.strptime(today_str, '%Y%m%d') - datetime.datetime.strptime(request_date, '%Y%m%d')).days >= 7:
        retri_num, fzd = binary_search(mail_count_num, request_date)
        print('开始从第%s封收取...' % retri_num)
        for i in range(retri_num, mail_count_num, 1):
            print(i)
            msg, address, subject, send_date = retrive(pop_conn, i)
            print(subject)
            if get_title_date(address, subject):
                title_date = get_title_date(address, subject)[0]
                if title_date == request_date:
                    print('下载邮件...')
                    download_attachment(msg)
                    print('下载完毕...')
                else:
                    skipped_num = skipped_num + 1
                    print('标题日期不匹配,跳过...')
            else:
                skipped_num = skipped_num + 1
                print('标题日期不存在,跳过...')
    else:
        for i in range(mail_count_num, 0, -1):
            msg, address, subject, send_date = retrive(pop_conn, i)
            print(subject)
            if get_title_date(address, subject):
                title_date = get_title_date(address, subject)[0]
                if title_date == request_date:
                    download_attachment(msg)
                else:
                    skipped_num = skipped_num + 1
                    print('标题日期不匹配,跳过...')
            else:
                skipped_num = skipped_num + 1
                print('标题日期不存在,跳过...')
    print('一共跳过%s封邮件' % skipped_num)

    wb.save('%s汇总表.xlsx' % foldername)
    print('===============任务完成===============\n')
    pop_conn.quit()
