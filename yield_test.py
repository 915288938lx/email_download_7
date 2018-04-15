import xlrd
from openpyxl import Workbook

# 查找表格所需单元格数据,并返回
str_list = ['日期', '单位净值', '市价', '银行存款', '存出保证金']



#返回str_list的坐标的迭代器
def find_cell_value_r_c(open_sheet, str_list):
    sheet = open_sheet
    rows = sheet.nrows  # 行数
    cols = sheet.ncols  # 列数
    # 每个要查找值的行数与列数
    for r in range(0, rows):
        for c in range(0, cols):
            for str_ in str_list:
                if str_ in sheet.cell(r, c).value:
                    yield r,c
def get_row_col_index(it):
    li = []
    for x in it:
        li.append(x)
    return li






if __name__ == '__main__':
    li_openpy = []
    wb = Workbook()
    active_sheet = wb.active
    open_sheet = xlrd.open_workbook('11.xls').sheet_by_index(0)
    it = find_cell_value_r_c(open_sheet, str_list)
    indx = get_row_col_index(it)
    excel_date = open_sheet.cell(indx[0][0],indx[0][1]).value

    excel_net_value = open_sheet.cell(indx[1][0],indx[1][1]).value

    excel_cash = open_sheet.cell(indx[3][0],indx[2][1]+1).value

    excel_secure_cash = open_sheet.cell(indx[4][0],indx[2][1]+1).value

    li_openpy.append(excel_date)
    li_openpy.append(excel_net_value)
    li_openpy.append(excel_cash)
    li_openpy.append(excel_secure_cash)
    active_sheet.append(li_openpy)
    wb.save('huizongbiao.xlsx')
# # #
#
# it = find_cell_value_r_c(open_sheet, str_list)
# li = []
# for x in it:
#     li.append(x)
# print(li)
# print(li[0][0])
#     #
#     #
#     #
#     # li = []
#     # row_index = get_row_index(it)
#     # col_index = get_col_index(it)
#     # wb = Workbook('huizongbiao.xlsx')
#     # active_sheet = wb.active
#     # excel_date = it[0,1]
#     #
#     # excel_net_value =
#     #
#     # excel_cash_net =
#     #
#     # excel_secure_cash =
